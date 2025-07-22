import pandas as pd
import openpyxl
import re
from datetime import datetime
import numpy as np
from io import BytesIO
import os

class VerificadorAnalyzer:
    def __init__(self):
        self.df = None
        self.df_original = None
        self.errores_originales = []
        self.errores_post_correccion = []
        self.archivo_original_path = None
        self.workbook_original = None
        self.archivo_temporal = None
        
    def corregir_nif(self, nif):
        """
        Corrige NIFs que tienen guiones mal colocados
        L-NNNNNNNN -> LNNNNNNNN
        NNNNNNNN-L -> NNNNNNNNL
        """
        if pd.isna(nif) or nif == "" or nif == 0:
            return nif, False, "No corregible"
        
        nif_str = str(nif).strip().upper()
        nif_original = nif_str
        
        # Eliminar guiones
        nif_str = nif_str.replace('-', '')
        
        # Verificar si era corregible y si ahora es válido
        if nif_original != nif_str:  # Había guiones
            if self.validar_nif_formato(nif_str):
                return nif_str, True, f"✅ {nif_original} → {nif_str}"
            else:
                return nif_original, False, "❌ Formato inválido incluso sin guiones"
        
        return nif_original, False, "Sin guiones para corregir"
    
    def validar_nif_formato(self, nif_str):
        """
        Valida solo el formato del NIF (sin mensaje de error)
        """
        # Patrón para LNNNNNNNN
        patron1 = re.compile(r'^[A-Z]\d{8}$')
        # Patrón para NNNNNNNNL
        patron2 = re.compile(r'^\d{8}[A-Z]$')
        
        return patron1.match(nif_str) or patron2.match(nif_str)
    
    def validar_nif(self, nif):
        """
        Valida el formato del NIF: LNNNNNNNN o NNNNNNNNL
        donde N es un número y L es una letra
        """
        if pd.isna(nif) or nif == "" or nif == 0:
            return False, "NIF vacío o nulo"
        
        nif_str = str(nif).strip().upper()
        
        if self.validar_nif_formato(nif_str):
            return True, "Válido"
        else:
            return False, f"Formato inválido: {nif_str}"
    
    def analizar_errores_originales(self, archivo_bytes, nombre_archivo):
        """
        Analiza el archivo original sin hacer correcciones
        """
        try:
            # Crear archivo temporal
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(archivo_bytes)
                temp_path = tmp_file.name
            
            # Guardar referencia al archivo temporal
            self.archivo_temporal = temp_path
            self.archivo_original_path = temp_path
            
            # Cargar el workbook original para preservar formato
            self.workbook_original = openpyxl.load_workbook(temp_path)
            
            # Leer el archivo saltando las primeras 6 filas y usando la fila 7 como encabezado
            self.df_original = pd.read_excel(temp_path, skiprows=6, header=0)
            self.df = self.df_original.copy()
            
            print(f"✅ Archivo cargado correctamente")
            print(f"📊 Dimensiones: {self.df.shape[0]} filas x {self.df.shape[1]} columnas")
            print(f"📋 Columnas detectadas: {list(self.df.columns)}")
            
            # Buscar columnas importantes
            col_verificador = None
            col_nif = None
            col_kg = None
            
            for col in self.df.columns:
                col_lower = str(col).lower()
                if 'verificador' in col_lower:
                    col_verificador = col
                elif 'nif' in col_lower and 'viticultor' in col_lower:
                    col_nif = col
                elif 'kg' in col_lower or 'kilo' in col_lower or 'peso' in col_lower:
                    col_kg = col
            
            print(f"\n🔍 Columnas identificadas:")
            print(f"   Verificador: {col_verificador}")
            print(f"   NIF Viticultor: {col_nif}")
            print(f"   Kg: {col_kg}")
            
            # Limpiar la lista de errores originales
            self.errores_originales = []
            
            # Analizar cada fila SIN hacer correcciones
            for index, row in self.df.iterrows():
                fila_errores = []
                correcciones_posibles = []
                
                # Verificar valores nulos o ceros en todas las columnas
                for col in self.df.columns:
                    valor = row[col]
                    if pd.isna(valor) or valor == 0 or valor == "":
                        if col == col_kg and valor == 0:
                            fila_errores.append(f"Campo '{col}' = 0 (se eliminará)")
                        else:
                            fila_errores.append(f"Campo '{col}' vacío/nulo/cero")
                
                # Validar NIF y verificar si es corregible
                if col_nif and col_nif in row:
                    es_valido, mensaje = self.validar_nif(row[col_nif])
                    if not es_valido:
                        # Verificar si se puede corregir
                        nif_corregido, es_corregible, detalle_correccion = self.corregir_nif(row[col_nif])
                        if es_corregible:
                            fila_errores.append(f"NIF: {mensaje} - CORREGIBLE")
                            correcciones_posibles.append(f"NIF: {detalle_correccion}")
                        else:
                            fila_errores.append(f"NIF: {mensaje} - NO CORREGIBLE")
                            correcciones_posibles.append(f"NIF: {detalle_correccion}")
                
                # Si hay errores, agregar la fila completa
                if fila_errores:
                    error_info = {
                        'Fila': index + 7 + 1,  # +7 por las filas saltadas, +1 por índice base 0
                        'Verificador': row.get(col_verificador, 'N/A'),
                        'Errores': '; '.join(fila_errores),
                        'Correcciones_Posibles': '; '.join(correcciones_posibles) if correcciones_posibles else 'Ninguna',
                        'Datos_Completos': row.to_dict(),
                        'Index_Original': index
                    }
                    self.errores_originales.append(error_info)
            
            # NO llamar a ninguna función de mostrar resultados aquí
            # Solo retornar True para indicar éxito
            return True
            
        except Exception as e:
            print(f"❌ Error al procesar el archivo: {str(e)}")
            return False
    
    def aplicar_correcciones(self):
        """
        Aplica las correcciones automáticas y analiza errores restantes
        """
        if self.df is None:
            return False
        
        # Buscar columnas importantes
        col_verificador = None
        col_nif = None
        col_kg = None
        
        for col in self.df.columns:
            col_lower = str(col).lower()
            if 'verificador' in col_lower:
                col_verificador = col
            elif 'nif' in col_lower and 'viticultor' in col_lower:
                col_nif = col
            elif 'kg' in col_lower or 'kilo' in col_lower or 'peso' in col_lower:
                col_kg = col
        
        # CORRECCIONES AUTOMÁTICAS
        nifs_corregidos = 0
        filas_eliminadas_kg = 0
        
        print(f"\n🔧 APLICANDO CORRECCIONES AUTOMÁTICAS...")
        print("="*50)
        
        # 1. Corregir NIFs con guiones
        if col_nif:
            print(f"🔧 Corrigiendo NIFs con guiones...")
            for index, row in self.df.iterrows():
                nif_corregido, fue_corregido, detalle = self.corregir_nif(row[col_nif])
                if fue_corregido:
                    self.df.at[index, col_nif] = nif_corregido
                    nifs_corregidos += 1
                    print(f"   ✅ Fila {index + 7 + 1}: {detalle}")
        
        # 2. Eliminar filas donde kg = 0
        if col_kg:
            print(f"\n🗑️ Eliminando filas con Kg = 0...")
            filas_eliminadas = []
            for index, row in self.df.iterrows():
                if row[col_kg] == 0:
                    filas_eliminadas.append(index + 7 + 1)
            
            self.df = self.df[self.df[col_kg] != 0]
            filas_eliminadas_kg = len(filas_eliminadas)
            
            if filas_eliminadas:
                print(f"   🗑️ Eliminadas {filas_eliminadas_kg} filas: {filas_eliminadas}")
            else:
                print(f"   ℹ️ No se encontraron filas con Kg = 0")
        
        # Resetear índices después de eliminar filas
        self.df = self.df.reset_index(drop=True)
        
        print(f"\n📊 RESUMEN DE CORRECCIONES:")
        print(f"   📝 NIFs corregidos: {nifs_corregidos}")
        print(f"   🗑️ Filas eliminadas (Kg=0): {filas_eliminadas_kg}")
        
        # Analizar errores restantes
        self.errores_post_correccion = []
        
        for index, row in self.df.iterrows():
            fila_errores = []
            
            # Verificar valores nulos o ceros en todas las columnas
            for col in self.df.columns:
                valor = row[col]
                if pd.isna(valor) or valor == 0 or valor == "":
                    fila_errores.append(f"Campo '{col}' vacío/nulo/cero")
            
            # Validaciones específicas
            if col_nif and col_nif in row:
                es_valido, mensaje = self.validar_nif(row[col_nif])
                if not es_valido:
                    fila_errores.append(f"NIF: {mensaje}")
            
            # Si hay errores, agregar la fila completa
            if fila_errores:
                error_info = {
                    'Fila': index + 7 + 1,  # +7 por las filas saltadas, +1 por índice base 0
                    'Verificador': row.get(col_verificador, 'N/A'),
                    'Errores': '; '.join(fila_errores),
                    'Datos_Completos': row.to_dict()
                }
                self.errores_post_correccion.append(error_info)
        
        return True
    
    def generar_archivo_corregido(self):
        """
        Genera el archivo Excel corregido manteniendo el formato original
        """
        try:
            if self.df is None or self.archivo_temporal is None:
                print("❌ No hay datos o archivo temporal disponible")
                return None
            
            # Cargar el workbook original nuevamente
            wb = openpyxl.load_workbook(self.archivo_temporal)
            ws = wb.active
            
            # Buscar columnas importantes en el DataFrame
            col_nif = None
            col_kg = None
            
            for col in self.df.columns:
                col_lower = str(col).lower()
                if 'nif' in col_lower and 'viticultor' in col_lower:
                    col_nif = col
                elif 'kg' in col_lower or 'kilo' in col_lower or 'peso' in col_lower:
                    col_kg = col
            
            print(f"🔍 Columnas para corrección:")
            print(f"   NIF: {col_nif}")
            print(f"   Kg: {col_kg}")
            
            # Leer el archivo original completo para obtener las filas eliminadas
            df_completo = pd.read_excel(self.archivo_temporal, skiprows=6, header=0)
            
            # Identificar qué filas fueron eliminadas (las que tenían kg = 0)
            filas_a_eliminar = []
            if col_kg:
                for index, row in df_completo.iterrows():
                    if row[col_kg] == 0:
                        fila_excel = index + 8  # +7 por skiprows, +1 por ser 1-indexado
                        filas_a_eliminar.append(fila_excel)
                        print(f"   📍 Fila a eliminar: {fila_excel} (Kg=0)")
            
            # Aplicar correcciones de NIFs al workbook
            if col_nif:
                col_index_nif = None
                # Buscar la columna de NIF en la fila de encabezados (fila 7)
                for col_idx in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=7, column=col_idx).value
                    if header_value and 'nif' in str(header_value).lower() and 'viticultor' in str(header_value).lower():
                        col_index_nif = col_idx
                        break
                
                print(f"   📍 Columna NIF en Excel: {col_index_nif}")
                
                if col_index_nif:
                    nifs_corregidos_excel = 0
                    for df_index, row in self.df_original.iterrows():
                        excel_row = df_index + 8  # +7 por skiprows, +1 por ser 1-indexado
                        
                        # Verificar si este NIF necesita corrección
                        nif_original = str(row[col_nif]).strip().upper()
                        nif_corregido, fue_corregido, detalle = self.corregir_nif(nif_original)
                        
                        if fue_corregido:
                            ws.cell(row=excel_row, column=col_index_nif, value=nif_corregido)
                            nifs_corregidos_excel += 1
                            print(f"   ✅ Excel fila {excel_row}: {nif_original} → {nif_corregido}")
                    
                    print(f"📝 Total NIFs corregidos en Excel: {nifs_corregidos_excel}")
            
            # Eliminar filas con kg = 0 (desde abajo hacia arriba para no afectar índices)
            if filas_a_eliminar:
                print(f"🗑️ Eliminando {len(filas_a_eliminar)} filas con Kg=0...")
                for fila_excel in sorted(filas_a_eliminar, reverse=True):
                    print(f"   🗑️ Eliminando fila {fila_excel}")
                    ws.delete_rows(fila_excel)
            
            # Guardar en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            print("✅ Archivo Excel generado exitosamente")
            return output.getvalue()
            
        except Exception as e:
            print(f"❌ Error detallado al generar archivo corregido: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def cleanup(self):
        """
        Limpia archivos temporales
        """
        if self.archivo_temporal and os.path.exists(self.archivo_temporal):
            try:
                os.remove(self.archivo_temporal)
                print("🗑️ Archivo temporal limpiado")
            except:
                pass